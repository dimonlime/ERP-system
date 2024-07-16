from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from database import Income, Payment, async_session_ODDS
from sqlalchemy.future import select


async def frame_report_ODDS(period, excel=r"ODDS/excel_files/report_ODDS.xlsx"):
    wb = Workbook()
    ws = wb.active
    period = period.split('-')
    year = period[0]
    month = period[1]

    async with async_session_ODDS() as session:
        result_incomes = await session.execute(
            select(Income).where(
                Income.date.startswith(f"{year}-{str(month).zfill(2)}")
            )
        )
        result_payments = await session.execute(
            select(Payment).where(
                Payment.date.startswith(f"{year}-{str(month).zfill(2)}")
            )
        )

        incomes = result_incomes.scalars().all()
        payments = result_payments.scalars().all()
    print(incomes[0])
    print(payments[0])
    print(len(payments))
    period = payments[0].date
    ws.title = 'Отчет'
    ws['A1'] = 'Наименование показателя'
    ws['B1'] = 'Код'
    ws['C1'] = f"За {period}"
    ws['A2'] = 'Денежные потоки от текущих операция'
    ws['A2'].font = Font(bold=True)
    ws.merge_cells('A2:C2')
    ws['A3'] = incomes[0].name
    ws['B3'] = incomes[0].code
    ws['C3'] = incomes[0].amount
    ws['A3'].font = Font(bold=True)
    row = 4
    for i in range(1, len(incomes) - 1):
        name = incomes[i].name
        code = incomes[i].code
        amount = incomes[i].amount
        ws[f"A{row}"] = name
        ws[f"B{row}"] = code
        ws[f"C{row}"] = amount
        row += 1

    ws[f"A{row}"] = payments[0].name
    ws[f"B{row}"] = payments[0].code
    ws[f"C{row}"] = payments[0].amount
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    for j in range(1, len(payments) - 2):
        name = payments[j].name
        code = payments[j].code
        amount = payments[j].amount
        ws[f"A{row}"] = name
        ws[f"B{row}"] = code
        ws[f"C{row}"] = amount
        row += 1
    ws[f"A{row}"] = payments[len(payments) - 1].name
    ws[f"B{row}"] = payments[len(payments) - 1].code
    ws[f"C{row}"] = payments[len(payments) - 1].amount
    ws.column_dimensions['A'].width = 65
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 25
    wb.save(excel)